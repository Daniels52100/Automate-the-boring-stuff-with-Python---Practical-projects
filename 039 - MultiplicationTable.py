#! python3
# This program creates a sheet with the multiplication table for the number passed as argument.
import sys, openpyxl
from openpyxl.styles import Font

# Creating spreadsheet:
wb = openpyxl.Workbook()
sheet = wb.active

# Reading the argument:
try:    
    dimensions = sys.argv[1]
except (IndexError, ValueError, TypeError):
    print('You must enter an integer as argument.')
    sys.exit()

# Setting up the headers:
font = Font(bold=True)
for i in range(2, dimensions + 2):
    for c in (sheet.cell(row=i, column=1), sheet.cell(row=1, column=i)):
        c.value = i-1
        c.font = font

# Defining the area in which the formula will be aplied:
start = 'B2'
end = sheet.cell(row=1 + dimensions, column=1 + dimensions).coordinate
# We have to add 1 to the coordinates because we skipped cell A1.

for row in sheet[start:end]:
    for cell in row:
        # Defining the multiplications' factors:
        factor1 = 'A' + str(cell.row)
        factor2 = cell.column + '1'
        cell.value = '=' + factor1 + '*' + factor2

# Saving sheet:
wb.save('C:\\Users\\Gabriel\\Documents\\UFU\\Cursos e Livros\\Livros\\Automatize Boring Stuff\\Practice Projects\\Planilhas para teste\\Tableof%s.xlsx' % (dimensions))