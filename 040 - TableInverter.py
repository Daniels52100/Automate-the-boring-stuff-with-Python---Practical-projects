# This program gets a spreadsheet and inverts its rows and columns.
import openpyxl, sys
from os.path import dirname, basename
# This is the path to the table that will be inverted. The inverted version will
# be created in the same folder.
path = '.\\Practice Projects\\Planilhas para teste\\teste.xlsx'

# Opening sheet and creating the inverted version:
wb = openpyxl.load_workbook(path)
sheet = wb.active
invWb = openpyxl.Workbook()
invSheet = invWb.active

# Inverting:
for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        invSheet.cell(row=j, column=i).value = sheet.cell(row=i, column=j).value
        # If there are other things to copy (such as font) we put them here with the logic above.
invWb.save((dirname(path) + '\\inverted_' + basename(path)))
