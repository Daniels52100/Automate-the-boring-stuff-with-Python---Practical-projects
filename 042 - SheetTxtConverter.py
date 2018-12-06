#! python3
# This program converts a folder of .txt files into a sheet or vice versa.
import openpyxl, os, sys
# If the user enters the arguments wrongly, the program displays a "how to use" message.
try:
    # Text to sheet
    if sys.argv[1] == 'TS':
        # Creating .xlsx file: 
        wb = openpyxl.Workbook()
        sheet = wb.active
        path = sys.argv[2] # folder containing text files
        os.chdir(path)
        col = 1
        # Searching for .txt files:
        for files in os.listdir(path):
            if files.endswith('.txt'):
                # Copying each line as a cell for the correspondant column.
                lines = open(files, 'r').readlines()
                for i in range(len(lines)):
                    sheet.cell(row=i + 1, column=col).value = lines[i]
                col += 1
        wb.save(path + '\\textsInSheet.xlsx')

    # Sheet to text:
    elif sys.argv[1] == 'ST':
        path = sys.argv[2]
        os.chdir(os.path.dirname(path))
        filename = os.path.basename(path)
        # Opening .xlsx:
        wb = openpyxl.load_workbook(os.path.join('.', filename))
        sheet = wb.active
        for col in sheet.iter_cols():
            # Creating a .txt file for each column:
            txtFile = open(os.path.join('.', 'textFromCol%s.txt' % col[0].column), 'w')
            for cell in col:
                # Writing down the cells to the .txt file:
                if cell.value != None:
                    txtFile.write(str(cell.value))
            txtFile.close()

    else:
        print('First argument should be "TS" (text to sheet) or "ST" (sheet to text).')

# Displaying "how to use" message:
except (FileNotFoundError, PermissionError, NotADirectoryError, AttributeError, TypeError, IndexError):
    print('Something went wrong. How to use:')
    print('    1st argument should be "ST" if you want sheet to text or "TS" if you want text to sheet.')
    print('    If you choose ST, you should pass a path to a .xlsx file as 2nd argument.')
    print('    If you choose TS, you should pass a path to a folder with the .txt files as 2nd argument.')
    sys.exit()
        
