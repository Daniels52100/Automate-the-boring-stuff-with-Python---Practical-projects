#! python3
# This program copies a .xlsx file to .csv format or vice versa.
import openpyxl, csv, os, sys
try:
    # Receiving and parsing argument:
    path = sys.argv[1]
    os.chdir(os.path.dirname(path))
    filename = os.path.basename(path)

    # 1st case: .xlsx file passed as argument - will be converted to .csv
    if filename.endswith('.xlsx'):
        xlFile = openpyxl.load_workbook(filename)
        xlSheet = xlFile.active
        csvFile = open(filename[:-4] + 'csv', 'w', newline='')
        csvWriter = csv.writer(csvFile)
        for row in xlSheet.iter_rows():
            csvRow = []
            for cell in row:
                csvRow.append(cell.value)
            csvWriter.writerow(csvRow)

    # 2nd case: .csv file passed as argument - will be converted to .xlsx
    elif filename.endswith('.csv'):
        csvFile = open(filename, 'r')
        csvReader = csv.reader(csvFile)
        xlFile = openpyxl.Workbook()
        xlSheet = xlFile.active
        for row in csvReader:
            for col in range(1, len(row) + 1):
                xlSheet.cell(row=csvReader.line_num, column=col).value = row[col - 1]
        xlFile.save(filename[:-3] + 'xlsx')

    # default:.
    else:
        print('You should enter a .xlsx or .csv file path as argument.')

except (IndexError, FileNotFoundError):
    print('Something went wrong. Enter a .csv or .xlsx file absolute path as argument.')
    print('If there are spaces in path, put the whole path between quotes.')